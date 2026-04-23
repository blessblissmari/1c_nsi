from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.models import (
    Characteristic,
    ClassCharacteristic,
    Document,
    EquipmentClass,
    EquipmentModel,
    EquipmentSubclass,
    TORCharacteristic,
    Unit,
)
from app.schemas.schemas import (
    BulkVerifyRequest,
    CharacteristicCreate,
    CharacteristicRead,
    ClassCharacteristicRead,
    MessageResponse,
    TORCharacteristicRead,
    TORCharacteristicUpdate,
    UnitCreate,
    UnitRead,
)
from app.services.ai_service import yandex_ai
from app.services.analogs import search_analogs_in_db

router = APIRouter(prefix="/mass-processing", tags=["Окно 3 — Массовая обработка моделей"])


def _ensure_unit(db: Session, unit_symbol: str) -> Unit:
    unit = db.query(Unit).filter(Unit.symbol == unit_symbol).first()
    if unit:
        return unit
    unit = Unit(name=unit_symbol, symbol=unit_symbol)
    db.add(unit)
    db.flush()
    return unit


def _ensure_characteristic(
    db: Session, name: str, unit_symbol: str | None, class_id: int | None, subclass_id: int | None
) -> Characteristic:
    unit_id = None
    if unit_symbol:
        unit_id = _ensure_unit(db, unit_symbol).id

    existing = (
        db.query(Characteristic)
        .filter(
            Characteristic.name == name,
            Characteristic.class_id == class_id,
            Characteristic.subclass_id == subclass_id,
        )
        .first()
    )
    if existing:
        if unit_id and existing.unit_id is None:
            existing.unit_id = unit_id
        return existing

    ch = Characteristic(name=name, unit_id=unit_id, class_id=class_id, subclass_id=subclass_id)
    db.add(ch)
    db.flush()
    return ch


def _upsert_tor_value(
    db: Session,
    model_id: int,
    characteristic: Characteristic,
    value: str,
    source_type: str,
    confidence: float,
    source_url: str | None = None,
) -> None:
    tv = (
        db.query(TORCharacteristic)
        .filter(
            TORCharacteristic.model_id == model_id,
            TORCharacteristic.characteristic_id == characteristic.id,
        )
        .first()
    )
    if not tv:
        tv = TORCharacteristic(model_id=model_id, characteristic_id=characteristic.id)
        db.add(tv)
    tv.value = value
    tv.source_type = source_type
    tv.confidence = confidence
    tv.source_url = source_url
    tv.verified = False


@router.post("/required-from-docs/{model_id}", response_model=MessageResponse)
def required_from_docs(model_id: int, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")
    if not model.class_id:
        raise HTTPException(400, "Model not classified")

    required_q = db.query(ClassCharacteristic).filter(ClassCharacteristic.class_id == model.class_id)
    if model.subclass_id:
        required_q = required_q.filter(
            (ClassCharacteristic.subclass_id == model.subclass_id)
            | (ClassCharacteristic.subclass_id.is_(None))
        )
    required = required_q.all()
    if not required:
        raise HTTPException(400, "No required characteristics for class/subclass")

    docs = db.query(Document).filter(Document.model_id == model.id).order_by(Document.priority.desc()).all()
    text = "\n\n".join([(d.parsed_content or "") for d in docs if (d.parsed_content or "").strip()])
    if not text.strip():
        raise HTTPException(400, "No parsed document text")

    names = [r.name for r in required]
    extracted = yandex_ai.extract_characteristics_from_text(text, names, require_units=True)

    filled = 0
    for item in extracted:
        ch_name = str(item.get("characteristic_name") or "").strip()
        val = str(item.get("value") or "").strip()
        if not ch_name or not val:
            continue
        expected = next((r for r in required if r.name == ch_name), None)
        unit_symbol = item.get("unit") or (expected.unit_symbol if expected else None)
        ch = _ensure_characteristic(db, ch_name, unit_symbol, model.class_id, model.subclass_id)
        _upsert_tor_value(db, model.id, ch, val, "vector_store", float(item.get("confidence") or 0.85))
        filled += 1

    db.commit()
    return MessageResponse(message=f"Из документов заполнено: {filled}")


@router.post("/required-from-web/{model_id}", response_model=MessageResponse)
def required_from_web(model_id: int, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")
    if not model.class_id:
        raise HTTPException(400, "Model not classified")

    required_q = db.query(ClassCharacteristic).filter(ClassCharacteristic.class_id == model.class_id)
    if model.subclass_id:
        required_q = required_q.filter(
            (ClassCharacteristic.subclass_id == model.subclass_id)
            | (ClassCharacteristic.subclass_id.is_(None))
        )
    required = required_q.all()
    if not required:
        raise HTTPException(400, "No required characteristics for class/subclass")

    cls = db.query(EquipmentClass).get(model.class_id)
    names = [r.name for r in required]
    enriched = yandex_ai.enrich_characteristics_via_web(
        model.normalized_name or model.original_name, cls.name if cls else None, names
    )

    filled = 0
    for item in enriched:
        ch_name = str(item.get("characteristic_name") or "").strip()
        val = str(item.get("value") or "").strip()
        if not ch_name or not val:
            continue
        expected = next((r for r in required if r.name == ch_name), None)
        unit_symbol = item.get("unit") or (expected.unit_symbol if expected else None)
        ch = _ensure_characteristic(db, ch_name, unit_symbol, model.class_id, model.subclass_id)
        _upsert_tor_value(
            db, model.id, ch, val, "yandex_web", float(item.get("confidence") or 0.7), item.get("source_url")
        )
        filled += 1

    db.commit()
    return MessageResponse(message=f"Из интернета заполнено: {filled}")


@router.post("/other-from-docs/{model_id}", response_model=MessageResponse)
def other_from_docs(model_id: int, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")

    docs = db.query(Document).filter(Document.model_id == model.id).order_by(Document.priority.desc()).all()
    text = "\n\n".join([(d.parsed_content or "") for d in docs if (d.parsed_content or "").strip()])
    if not text.strip():
        raise HTTPException(400, "No parsed document text")

    required_names = []
    if model.class_id:
        required_q = db.query(ClassCharacteristic).filter(ClassCharacteristic.class_id == model.class_id)
        if model.subclass_id:
            required_q = required_q.filter(
                (ClassCharacteristic.subclass_id == model.subclass_id)
                | (ClassCharacteristic.subclass_id.is_(None))
            )
        required_names = [r.name for r in required_q.all()]

    extracted = yandex_ai.extract_other_characteristics_from_text(
        text, exclude_names=required_names, limit=20
    )
    filled = 0
    for item in extracted:
        ch_name = str(item.get("characteristic_name") or "").strip()
        val = str(item.get("value") or "").strip()
        if not ch_name or not val:
            continue
        unit_symbol = item.get("unit")
        ch = _ensure_characteristic(db, ch_name, unit_symbol, model.class_id, model.subclass_id)
        _upsert_tor_value(db, model.id, ch, val, "vector_store", float(item.get("confidence") or 0.75))
        filled += 1

    db.commit()
    return MessageResponse(message=f"Прочие из документов добавлены/обновлены: {filled}")


@router.get("/class-characteristics", response_model=list[ClassCharacteristicRead])
def get_class_characteristics(
    class_id: int | None = None,
    subclass_id: int | None = None,
    db: Session = Depends(get_db),
):
    q = db.query(ClassCharacteristic)
    if class_id:
        q = q.filter(ClassCharacteristic.class_id == class_id)
    if subclass_id is not None:
        q = q.filter(
            (ClassCharacteristic.subclass_id == subclass_id) | (ClassCharacteristic.subclass_id.is_(None))
        )
    return q.all()


@router.post("/upload-class-characteristics", response_model=MessageResponse)
async def upload_class_characteristics(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Upload class/subclass -> characteristics catalog from xlsx like:
    columns: Класс, Подкласс, Характеристика N, Ед.измерения N ...
    """
    import os
    import tempfile

    import openpyxl

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb.active

        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        header = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
        col_map = {name: idx for idx, name in enumerate(header)}
        if "Класс" not in col_map:
            raise HTTPException(status_code=400, detail="Не найдена колонка 'Класс'")

        def is_blue_fill(cell) -> bool:
            try:
                fill = getattr(cell, "fill", None)
                if not fill or not getattr(fill, "patternType", None):
                    return False
                fg = getattr(fill, "fgColor", None)
                rgb = getattr(fg, "rgb", None)
                if not rgb:
                    return False
                rgb = str(rgb).upper()
                common_blues = {
                    "FF00B0F0",
                    "FF0070C0",
                    "FF5B9BD5",
                    "FF2F75B5",
                    "FF4F81BD",
                    "FF1F4E79",
                }
                return rgb in common_blues
            except Exception:
                return False

        created = 0
        updated = 0
        skipped = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=False):
            row_vals = [c.value for c in row_cells]
            cls_name = row_vals[col_map.get("Класс")]
            sub_name = row_vals[col_map.get("Подкласс")] if "Подкласс" in col_map else None
            if not cls_name:
                continue

            cls = db.query(EquipmentClass).filter(EquipmentClass.name == str(cls_name).strip()).first()
            if not cls:
                skipped += 1
                continue

            subclass_id = None
            if sub_name and str(sub_name).strip():
                sub = (
                    db.query(EquipmentSubclass)
                    .filter(
                        EquipmentSubclass.class_id == cls.id,
                        EquipmentSubclass.name == str(sub_name).strip(),
                    )
                    .first()
                )
                if sub:
                    subclass_id = sub.id

            for i in range(1, 30):
                ch_col = f"Характеристика {i}"
                u_col = f"Ед.измерения {i}"
                if ch_col not in col_map:
                    break
                ch_cell = row_cells[col_map[ch_col]]
                ch = ch_cell.value
                if not ch:
                    continue
                ch_name = str(ch).strip()
                unit_symbol = None
                if u_col in col_map:
                    u = row_vals[col_map[u_col]]
                    if u is not None and str(u).strip():
                        unit_symbol = str(u).strip()

                required = is_blue_fill(ch_cell)

                existing = (
                    db.query(ClassCharacteristic)
                    .filter(
                        ClassCharacteristic.class_id == cls.id,
                        ClassCharacteristic.subclass_id == subclass_id,
                        ClassCharacteristic.name == ch_name,
                    )
                    .first()
                )
                if existing:
                    if existing.unit_symbol != unit_symbol or existing.required != required:
                        existing.unit_symbol = unit_symbol
                        existing.required = required
                        updated += 1
                else:
                    db.add(
                        ClassCharacteristic(
                            class_id=cls.id,
                            subclass_id=subclass_id,
                            name=ch_name,
                            unit_symbol=unit_symbol,
                            required=required,
                        )
                    )
                    created += 1

        db.commit()
        return MessageResponse(
            message=f"Загружено характеристик: {created} (обновлено: {updated}, пропущено строк: {skipped})"
        )
    finally:
        os.unlink(tmp_path)


@router.get("/characteristics", response_model=list[CharacteristicRead])
def get_characteristics(
    class_id: int | None = None,
    subclass_id: int | None = None,
    db: Session = Depends(get_db),
):
    q = db.query(Characteristic)
    if class_id:
        q = q.filter((Characteristic.class_id == class_id) | (Characteristic.class_id.is_(None)))
    if subclass_id:
        q = q.filter((Characteristic.subclass_id == subclass_id) | (Characteristic.subclass_id.is_(None)))
    chars = q.all()
    result = []
    for c in chars:
        result.append(
            CharacteristicRead(
                id=c.id,
                name=c.name,
                unit_id=c.unit_id,
                class_id=c.class_id,
                subclass_id=c.subclass_id,
                unit_symbol=c.unit.symbol if c.unit else None,
            )
        )
    return result


@router.post("/characteristics", response_model=CharacteristicRead)
def create_characteristic(data: CharacteristicCreate, db: Session = Depends(get_db)):
    char = Characteristic(**data.model_dump())
    db.add(char)
    db.commit()
    db.refresh(char)
    return CharacteristicRead(
        id=char.id,
        name=char.name,
        unit_id=char.unit_id,
        class_id=char.class_id,
        subclass_id=char.subclass_id,
        unit_symbol=char.unit.symbol if char.unit else None,
    )


@router.post("/upload-characteristics", response_model=MessageResponse)
async def upload_characteristics(file: UploadFile = File(...), db: Session = Depends(get_db)):
    import os
    import tempfile

    from app.services.file_parser import parse_xlsx

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        rows = parse_xlsx(tmp_path)
        created_chars = 0
        created_units = 0

        for row in rows:
            name = (
                row.get("Характеристика")
                or row.get("Наименование")
                or row.get("characteristic")
                or row.get("name")
            )
            unit_name = (
                row.get("Ед.изм.") or row.get("Единица") or row.get("unit") or row.get("Ед. измерения")
            )
            unit_symbol = row.get("Обозначение") or row.get("symbol") or row.get("Символ")
            class_name = row.get("Класс") or row.get("class")
            subclass_name = row.get("Подкласс") or row.get("subclass")

            if not name:
                continue

            unit_id = None
            if unit_name or unit_symbol:
                unit = db.query(Unit).filter(Unit.name == str(unit_name or unit_symbol)).first()
                if not unit:
                    unit = Unit(
                        name=str(unit_name or unit_symbol),
                        symbol=str(unit_symbol or unit_name),
                    )
                    db.add(unit)
                    db.flush()
                    created_units += 1
                unit_id = unit.id

            class_id = None
            subclass_id = None
            if class_name:
                cls = db.query(EquipmentClass).filter(EquipmentClass.name == str(class_name)).first()
                if cls:
                    class_id = cls.id
                    if subclass_name:
                        sub = (
                            db.query(EquipmentSubclass)
                            .filter(
                                EquipmentSubclass.name == str(subclass_name),
                                EquipmentSubclass.class_id == cls.id,
                            )
                            .first()
                        )
                        if sub:
                            subclass_id = sub.id

            existing = (
                db.query(Characteristic)
                .filter(
                    Characteristic.name == str(name),
                    Characteristic.class_id == class_id,
                )
                .first()
            )
            if not existing:
                char = Characteristic(
                    name=str(name),
                    unit_id=unit_id,
                    class_id=class_id,
                    subclass_id=subclass_id,
                )
                db.add(char)
                created_chars += 1

        db.commit()
        return MessageResponse(message=f"Loaded {created_chars} characteristics, {created_units} units")
    finally:
        os.unlink(tmp_path)


@router.get("/units", response_model=list[UnitRead])
def get_units(db: Session = Depends(get_db)):
    return db.query(Unit).all()


@router.post("/units", response_model=UnitRead)
def create_unit(data: UnitCreate, db: Session = Depends(get_db)):
    unit = Unit(**data.model_dump())
    db.add(unit)
    db.commit()
    db.refresh(unit)
    return unit


@router.post("/upload-units", response_model=MessageResponse)
async def upload_units(file: UploadFile = File(...), db: Session = Depends(get_db)):
    import os
    import tempfile

    from app.services.file_parser import parse_xlsx

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        rows = parse_xlsx(tmp_path)
        created = 0
        for row in rows:
            name = row.get("Наименование") or row.get("name") or row.get("Ед.изм.")
            symbol = row.get("Обозначение") or row.get("symbol") or row.get("Символ") or name

            if not name:
                continue

            existing = db.query(Unit).filter(Unit.name == str(name)).first()
            if not existing:
                unit = Unit(name=str(name), symbol=str(symbol))
                db.add(unit)
                created += 1

        db.commit()
        return MessageResponse(message=f"Loaded {created} units")
    finally:
        os.unlink(tmp_path)


@router.post("/bind-characteristics/{model_id}", response_model=MessageResponse)
def bind_characteristics_to_tor(model_id: int, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")

    chars = (
        db.query(Characteristic)
        .filter((Characteristic.class_id == model.class_id) | (Characteristic.class_id.is_(None)))
        .all()
    )

    bound = 0
    for char in chars:
        existing = (
            db.query(TORCharacteristic)
            .filter(
                TORCharacteristic.model_id == model_id,
                TORCharacteristic.characteristic_id == char.id,
            )
            .first()
        )
        if not existing:
            tor_char = TORCharacteristic(model_id=model_id, characteristic_id=char.id)
            db.add(tor_char)
            bound += 1

    db.commit()
    return MessageResponse(message=f"Bound {bound} characteristics to TOR")


@router.get("/tor-characteristics/{model_id}", response_model=list[TORCharacteristicRead])
def get_tor_characteristics(model_id: int, db: Session = Depends(get_db)):
    items = db.query(TORCharacteristic).filter(TORCharacteristic.model_id == model_id).all()
    result = []
    for item in items:
        result.append(
            TORCharacteristicRead(
                id=item.id,
                model_id=item.model_id,
                characteristic_id=item.characteristic_id,
                value=item.value,
                source_type=item.source_type,
                confidence=item.confidence,
                source_url=item.source_url,
                verified=item.verified,
                characteristic_name=item.characteristic.name if item.characteristic else None,
                unit_symbol=item.characteristic.unit.symbol
                if item.characteristic and item.characteristic.unit
                else None,
            )
        )
    return result


@router.put("/tor-characteristics/{item_id}", response_model=TORCharacteristicRead)
def update_tor_characteristic(item_id: int, data: TORCharacteristicUpdate, db: Session = Depends(get_db)):
    item = db.query(TORCharacteristic).get(item_id)
    if not item:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    db.commit()
    db.refresh(item)
    return item


@router.post("/fill-characteristics-from-source/{model_id}", response_model=MessageResponse)
def fill_characteristics_from_source(model_id: int, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")

    tor_chars = (
        db.query(TORCharacteristic)
        .filter(
            TORCharacteristic.model_id == model_id,
            TORCharacteristic.value.is_(None),
        )
        .all()
    )

    if not tor_chars:
        return MessageResponse(message="No empty characteristics to fill")

    char_names = [tc.characteristic.name for tc in tor_chars if tc.characteristic]
    ai_results = yandex_ai.enrich_characteristics_via_vector_store(
        model.normalized_name or model.original_name, char_names
    )

    filled = 0
    for result in ai_results:
        char_name = result.get("characteristic_name")
        tor_char = next(
            (tc for tc in tor_chars if tc.characteristic and tc.characteristic.name == char_name), None
        )
        if tor_char:
            tor_char.value = result.get("value")
            tor_char.source_type = result.get("source", "vector_store")
            tor_char.confidence = result.get("confidence", 0.85)
            filled += 1

    db.commit()
    return MessageResponse(message=f"Filled {filled} characteristics from Vector Store")


@router.post("/enrich-characteristics-from-web/{model_id}", response_model=MessageResponse)
def enrich_characteristics_from_web(model_id: int, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")

    tor_chars = (
        db.query(TORCharacteristic)
        .filter(
            TORCharacteristic.model_id == model_id,
            TORCharacteristic.value.is_(None),
        )
        .all()
    )

    if not tor_chars:
        return MessageResponse(message="No empty characteristics to enrich")

    char_names = [tc.characteristic.name for tc in tor_chars if tc.characteristic]
    class_name = model.eq_class.name if model.eq_class else None

    ai_results = yandex_ai.enrich_characteristics_via_web(
        model.normalized_name or model.original_name, class_name, char_names
    )

    filled = 0
    for result in ai_results:
        char_name = result.get("characteristic_name")
        tor_char = next(
            (tc for tc in tor_chars if tc.characteristic and tc.characteristic.name == char_name), None
        )
        if tor_char:
            tor_char.value = result.get("value")
            tor_char.source_type = result.get("source", "yandex_web")
            tor_char.confidence = result.get("confidence", 0.7)
            filled += 1

    db.commit()
    return MessageResponse(message=f"Enriched {filled} characteristics from web")


@router.post("/search-analogs/{model_id}")
def search_analogs(model_id: int, selected_chars: list[int] | None = None, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")

    characteristics = None
    if selected_chars:
        tor_chars = (
            db.query(TORCharacteristic)
            .filter(
                TORCharacteristic.model_id == model_id,
                TORCharacteristic.characteristic_id.in_(selected_chars),
                TORCharacteristic.value.isnot(None),
            )
            .all()
        )
        characteristics = {tc.characteristic.name: tc.value for tc in tor_chars if tc.characteristic}

    # Per TЗ "Оценка качества данных": do not invent analogs. Return only DB-backed results.
    return search_analogs_in_db(db=db, base_model=model, characteristics=characteristics)


@router.post("/verify", response_model=MessageResponse)
def bulk_verify(data: BulkVerifyRequest, db: Session = Depends(get_db)):
    db.query(TORCharacteristic).filter(TORCharacteristic.id.in_(data.ids)).update(
        {TORCharacteristic.verified: data.verified}, synchronize_session=False
    )
    db.commit()
    return MessageResponse(message=f"Updated {len(data.ids)} items")
