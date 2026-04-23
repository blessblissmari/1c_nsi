from __future__ import annotations

from datetime import datetime
import csv
import io
import os
import tempfile

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.models import EquipmentModel, ReliabilityMetric, FailureEvent
from app.schemas.schemas import (
    ReliabilityMetricCreate,
    ReliabilityMetricRead,
    ReliabilityMetricUpdate,
    BulkVerifyRequest,
    MessageResponse,
    FailureEventRead,
)
from app.services.ai_service import yandex_ai

router = APIRouter(prefix="/reliability", tags=["Окно 7 — Параметры надёжности"])


@router.get("/metrics", response_model=list[ReliabilityMetricRead])
def get_metrics(model_id: int, db: Session = Depends(get_db)):
    return db.query(ReliabilityMetric).filter(ReliabilityMetric.model_id == model_id).all()


@router.post("/metrics", response_model=ReliabilityMetricRead)
def create_metric(data: ReliabilityMetricCreate, db: Session = Depends(get_db)):
    metric = ReliabilityMetric(**data.model_dump())
    db.add(metric)
    db.commit()
    db.refresh(metric)
    return metric


@router.put("/metrics/{metric_id}", response_model=ReliabilityMetricRead)
def update_metric(metric_id: int, data: ReliabilityMetricUpdate, db: Session = Depends(get_db)):
    metric = db.query(ReliabilityMetric).get(metric_id)
    if not metric:
        raise HTTPException(404, "Metric not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(metric, k, v)
    db.commit()
    db.refresh(metric)
    return metric


@router.delete("/metrics/{metric_id}", response_model=MessageResponse)
def delete_metric(metric_id: int, db: Session = Depends(get_db)):
    metric = db.query(ReliabilityMetric).get(metric_id)
    if not metric:
        raise HTTPException(404, "Metric not found")
    db.delete(metric)
    db.commit()
    return MessageResponse(message="Deleted")


@router.post("/fill-from-source/{model_id}", response_model=MessageResponse)
def fill_from_source(model_id: int, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")

    ai_results = yandex_ai.enrich_reliability_via_vector_store(
        model.normalized_name or model.original_name
    )

    created = 0
    for result in ai_results:
        metric = ReliabilityMetric(
            model_id=model_id,
            metric_type=result.get("metric_type", ""),
            value=result.get("value"),
            unit=result.get("unit"),
            description=result.get("description"),
            source_type=result.get("source", "vector_store"),
            confidence=result.get("confidence", 0.85),
        )
        db.add(metric)
        created += 1

    db.commit()
    return MessageResponse(message=f"Created {created} metrics from Vector Store")


@router.post("/enrich-from-web/{model_id}", response_model=MessageResponse)
def enrich_from_web(model_id: int, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")

    class_name = model.eq_class.name if model.eq_class else None
    ai_results = yandex_ai.enrich_reliability_via_web(
        model.normalized_name or model.original_name, class_name
    )

    created = 0
    for result in ai_results:
        existing = db.query(ReliabilityMetric).filter(
            ReliabilityMetric.model_id == model_id,
            ReliabilityMetric.metric_type == result.get("metric_type"),
        ).first()

        if not existing:
            metric = ReliabilityMetric(
                model_id=model_id,
                metric_type=result.get("metric_type", ""),
                value=result.get("value"),
                unit=result.get("unit"),
                description=result.get("description"),
                source_type=result.get("source", "yandex_web"),
                confidence=result.get("confidence", 0.7),
                source_url=result.get("source_url"),
            )
            db.add(metric)
            created += 1

    db.commit()
    return MessageResponse(message=f"Enriched {created} metrics from web")


@router.post("/verify", response_model=MessageResponse)
def bulk_verify(data: BulkVerifyRequest, db: Session = Depends(get_db)):
    db.query(ReliabilityMetric).filter(ReliabilityMetric.id.in_(data.ids)).update(
        {ReliabilityMetric.verified: data.verified}, synchronize_session=False
    )
    db.commit()
    return MessageResponse(message=f"Updated {len(data.ids)} items")


@router.get("/failures", response_model=list[FailureEventRead])
def get_failures(model_id: int, db: Session = Depends(get_db)):
    return (
        db.query(FailureEvent)
        .filter(FailureEvent.model_id == model_id)
        .order_by(FailureEvent.occurred_at.asc().nulls_last(), FailureEvent.id.asc())
        .all()
    )


def _normalize_header(s: str) -> str:
    return "".join(
        ch.lower() for ch in s.strip() if ch.isalnum() or ch in ["_", "-", " "]
    ).replace(" ", "")


def _parse_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    txt = str(v).strip()
    if not txt:
        return None
    txt = txt.replace(",", ".")
    try:
        return float(txt)
    except Exception:
        return None


def _find_model_id(db: Session, model_text: str) -> int | None:
    if not model_text:
        return None
    raw = str(model_text).strip()
    if not raw:
        return None
    upper = raw.upper()

    m = (
        db.query(EquipmentModel)
        .filter((EquipmentModel.normalized_name == upper) | (EquipmentModel.original_name == raw))
        .first()
    )
    if m:
        return m.id

    m2 = db.query(EquipmentModel).filter(EquipmentModel.original_name.ilike(raw)).first()
    return m2.id if m2 else None


@router.post("/upload-failures", response_model=MessageResponse)
async def upload_failures(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Upload failures statistics (xlsx/csv).

    Expected columns (any order, RU/EN variants):
    - model / модель / код / наименование / обозначение (required)
    - occurred_at / дата / датаотказа (optional)
    - runtime_hours / наработка / наработкачасов (optional)
    - description / описание / комментарий (optional)
    """
    created = 0
    skipped = 0

    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        content = await file.read()
        text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            keys = {_normalize_header(k): k for k in (row.keys() or []) if k}
            model_key = (
                keys.get("model")
                or keys.get("модель")
                or keys.get("код")
                or keys.get("modelname")
                or keys.get("наименование")
                or keys.get("обозначение")
            )
            if not model_key:
                skipped += 1
                continue
            model_id = _find_model_id(db, row.get(model_key))
            if not model_id:
                skipped += 1
                continue

            occurred_at = None
            for hk in ["occurredat", "date", "дата", "датаотказа", "failuredate"]:
                src = keys.get(hk)
                if not src:
                    continue
                raw = row.get(src)
                if not raw:
                    break
                try:
                    occurred_at = datetime.fromisoformat(str(raw).strip())
                except Exception:
                    occurred_at = None
                break

            runtime_hours = None
            for hk in ["runtimehours", "runtime", "наработка", "наработкач", "наработкачасов"]:
                src = keys.get(hk)
                if not src:
                    continue
                runtime_hours = _parse_float(row.get(src))
                break

            desc_key = keys.get("description") or keys.get("описание") or keys.get("комментарий")
            description = (
                str(row.get(desc_key)).strip()
                if desc_key and row.get(desc_key) is not None and str(row.get(desc_key)).strip()
                else None
            )

            db.add(
                FailureEvent(
                    model_id=model_id,
                    occurred_at=occurred_at,
                    description=description,
                    runtime_hours=runtime_hours,
                    source_type="upload_csv",
                )
            )
            created += 1

        db.commit()
        return MessageResponse(
            message=f"Загружено отказов: {created} (пропущено строк: {skipped})"
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        header_vals = [str(c.value).strip() if c.value is not None else "" for c in header_row]
        header_norm = [_normalize_header(v) for v in header_vals]
        idx = {header_norm[i]: i for i in range(len(header_norm)) if header_norm[i]}

        def pick(*names: str) -> int | None:
            for n in names:
                if n in idx:
                    return idx[n]
            return None

        model_i = pick("model", "модель", "код", "modelname", "наименование", "обозначение")
        if model_i is None:
            raise HTTPException(status_code=400, detail="Не найдена колонка 'Модель/Model'")

        date_i = pick("occurredat", "date", "дата", "датаотказа", "failuredate")
        runtime_i = pick("runtimehours", "runtime", "наработка", "наработкач", "наработкачасов")
        desc_i = pick("description", "описание", "комментарий")

        for row in ws.iter_rows(min_row=2, values_only=True):
            model_text = row[model_i] if model_i < len(row) else None
            model_id = _find_model_id(db, model_text)
            if not model_id:
                skipped += 1
                continue

            occurred_at = None
            if date_i is not None and date_i < len(row) and row[date_i] is not None:
                v = row[date_i]
                if isinstance(v, datetime):
                    occurred_at = v
                else:
                    try:
                        occurred_at = datetime.fromisoformat(str(v).strip())
                    except Exception:
                        occurred_at = None

            runtime_hours = None
            if runtime_i is not None and runtime_i < len(row):
                runtime_hours = _parse_float(row[runtime_i])

            description = None
            if desc_i is not None and desc_i < len(row) and row[desc_i] is not None:
                description = str(row[desc_i]).strip()

            db.add(
                FailureEvent(
                    model_id=model_id,
                    occurred_at=occurred_at,
                    description=description,
                    runtime_hours=runtime_hours,
                    source_type="upload_xlsx",
                )
            )
            created += 1

        db.commit()
        return MessageResponse(
            message=f"Загружено отказов: {created} (пропущено строк: {skipped})"
        )
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


@router.post("/recalc-mtbf/{model_id}", response_model=MessageResponse)
def recalc_mtbf(model_id: int, db: Session = Depends(get_db)):
    model = db.query(EquipmentModel).get(model_id)
    if not model:
        raise HTTPException(404, "Model not found")

    failures = db.query(FailureEvent).filter(FailureEvent.model_id == model_id).all()
    if not failures:
        raise HTTPException(400, "No failure events for this model")

    with_runtime = [f for f in failures if f.runtime_hours is not None and f.runtime_hours > 0]
    mtbf_hours: float | None = None

    if with_runtime:
        mtbf_hours = sum(f.runtime_hours for f in with_runtime) / float(len(with_runtime))
    else:
        dated = sorted([f for f in failures if f.occurred_at is not None], key=lambda x: x.occurred_at)
        if len(dated) >= 2:
            deltas = []
            for i in range(1, len(dated)):
                dt = (dated[i].occurred_at - dated[i - 1].occurred_at).total_seconds() / 3600.0
                if dt > 0:
                    deltas.append(dt)
            if deltas:
                mtbf_hours = sum(deltas) / float(len(deltas))

    if mtbf_hours is None:
        raise HTTPException(
            400,
            "Insufficient data to calculate MTBF (need runtime_hours or at least 2 dates)",
        )

    description = "MTBF рассчитан из статистики отказов (загрузка)"
    existing = (
        db.query(ReliabilityMetric)
        .filter(ReliabilityMetric.model_id == model_id, ReliabilityMetric.metric_type == "mtbf")
        .first()
    )
    if existing:
        existing.value = mtbf_hours
        existing.unit = "ч"
        existing.description = description
        existing.source_type = "failures_upload"
        existing.confidence = 1.0
        existing.source_url = None
        existing.verified = False
    else:
        db.add(
            ReliabilityMetric(
                model_id=model_id,
                metric_type="mtbf",
                value=mtbf_hours,
                unit="ч",
                description=description,
                source_type="failures_upload",
                confidence=1.0,
                verified=False,
            )
        )

    db.commit()
    return MessageResponse(message=f"MTBF = {mtbf_hours:.2f} ч (обновлено)")

